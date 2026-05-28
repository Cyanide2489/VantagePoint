#!/usr/bin/env python3
"""Extract the SOC-CMM v2.4.2 maturity assessment from the official Excel
workbooks into a structured English JSON consumed by the Nuxt frontend.

Sources (in repo root):
  - 62-soc-cmm-242-basic.xlsx        question bank, guidance scenarios, scoring
  - 6-soc-cmm-23-nist-csf-mapping.xlsx   NIST CSF 2.0 descriptions

Output:
  - app/data/soc-cmm.en.json

We keep the MATURITY dimension (type 'M' in the _Output sheet) because those are
the questions whose 1-5 levels each carry a scenario description -- the core of
this self-assessment. Capability ('C') coverage questions are intentionally
omitted to keep the tool operable for non-specialists.
"""
import json
import re
import warnings
from collections import defaultdict
from pathlib import Path

import openpyxl

warnings.simplefilter("ignore")

ROOT = Path(__file__).resolve().parent.parent
BASIC = ROOT / "62-soc-cmm-242-basic.xlsx"
NISTF = ROOT / "6-soc-cmm-23-nist-csf-mapping.xlsx"
OUT = ROOT / "app" / "data" / "soc-cmm.en.json"

# domain sheet prefix -> (letter, english title). Order defines display order.
DOMAINS = [
    ("Business", "B", "Business"),
    ("People", "P", "People"),
    ("Process", "M", "Process"),
    ("Technology", "T", "Technology"),
    ("Services", "S", "Services"),
]

# the SOC-CMM "Detailed" maturity scale (shared label set for 1-5 questions)
DETAILED_LABELS = {1: "No", 2: "Partially", 3: "Averagely", 4: "Mostly", 5: "Fully"}

NUM_RE = re.compile(r"^\d+(\.\d+)*$")          # e.g. 1, 1.1, 2.1.3
CODE_RE = re.compile(r"^[A-Z] \d")             # e.g. "B 1.1"
NIST_RE = re.compile(r"[A-Z]{2}\.[A-Z]{2}-\d{2}")  # e.g. GV.OC-04


def load(path):
    return openpyxl.load_workbook(path, data_only=True)


def parse_guidance(wb):
    """code -> list[(value, scenario)] for values 1..5."""
    ws = wb["_Guidance"]
    cur = None
    out = defaultdict(list)
    for a, b, c in ws.iter_rows(min_row=1, max_col=3, values_only=True):
        if a and CODE_RE.match(str(a).strip()):
            code = re.sub(r"\s+", " ", str(a).strip())
            # the source _Guidance sheet has a duplicate 'S 3.12' header whose
            # second block belongs to a different question; keep the first only
            cur = None if out.get(code) else code
        elif cur and b is not None and str(b).strip().isdigit():
            val = int(str(b).strip())
            if 1 <= val <= 5:
                scenario = (str(c).strip() if c is not None else "")
                out[cur].append((val, scenario))
    return out


def parse_output(wb):
    """code -> {importance, nist:[...], in_scope} for MATURITY (type 'M') rows.

    Deduplicates repeated codes; merges NIST references across duplicate rows.
    """
    ws = wb["_Output"]
    out = {}
    for r in ws.iter_rows(min_row=2, max_col=12, values_only=True):
        code = r[0]
        if not code or not CODE_RE.match(str(code)) or r[2] != "M":
            continue
        code = re.sub(r"\s+", " ", str(code).strip())
        nist = NIST_RE.findall(str(r[5])) if r[5] else []
        rec = out.setdefault(
            code, {"importance": r[4] or 3, "in_scope": bool(r[1]), "nist": []}
        )
        for n in nist:
            if n not in rec["nist"]:
                rec["nist"].append(n)
    return out


def parse_domain_sheets(wb):
    """Walk each domain sheet to recover hierarchy + question text.

    Returns:
      texts: code -> question text
      sections: (letter, secNo) -> section title
      cat_titles: category-code (e.g. 'T 2.1') -> category title
    """
    texts = {}
    sections = {}
    for sheet in wb.sheetnames:
        if " - " not in sheet:
            continue
        dom = sheet.split(" - ")[0]
        match = next((d for d in DOMAINS if d[0] == dom), None)
        if not match:
            continue
        letter = match[1]
        ws = wb[sheet]
        sec_no = None
        for row in ws.iter_rows(min_row=8, max_col=3, values_only=True):
            c1, c2, c3 = row[0], row[1], row[2]
            # section header: col1 holds the (global) section number
            if c1 is not None and str(c1).strip().isdigit():
                sec_no = str(c1).strip()
                sections[(letter, sec_no)] = str(c2).strip() if c2 else ""
                continue
            if c2 is None:
                continue
            num = str(c2).strip()
            if not NUM_RE.match(num):
                continue  # 'Maturity' / 'Capability' / comment headers
            code = f"{letter} {num}"
            text = str(c3).strip() if c3 else ""
            if text:
                texts[code] = text
    return texts, sections


def parse_nist(wb):
    """subcode -> {function, category, subcategory} from NIST CSF 2.0 sheet."""
    ws = wb["NIST CSF 2.0"]
    out = {}
    cur_func = cur_cat = ""
    for r in ws.iter_rows(min_row=3, max_col=3, values_only=True):
        func, cat, sub = r[0], r[1], r[2]
        if func:
            cur_func = str(func).split("(")[0].strip()
        if cat and not sub:
            cur_cat = str(cat).split("(")[0].strip()
        if sub:
            m = NIST_RE.search(str(sub))
            if m:
                code = m.group(0)
                desc = str(sub).split(":", 1)[1].strip() if ":" in str(sub) else str(sub).strip()
                out[code] = {
                    "function": cur_func,
                    "category": cur_cat,
                    "subcategory": desc,
                }
    return out


def category_of(code):
    """For 'T 2.1.3' -> 'T 2.1'; for 'B 1.1' -> None (sits under section)."""
    letter, num = code.split(" ", 1)
    parts = num.split(".")
    if len(parts) >= 3:
        return f"{letter} {'.'.join(parts[:2])}"
    return None


def main():
    wb = load(BASIC)
    nwb = load(NISTF)

    guidance = parse_guidance(wb)
    output = parse_output(wb)
    texts, sections = parse_domain_sheets(wb)
    nist_desc = parse_nist(nwb)

    # assemble: keep M-type questions that have text + 1-5 scenarios
    domains = []
    nist_used = {}
    kept = skipped = 0

    for dom_name, letter, dom_title in DOMAINS:
        dom_secs = {}
        # iterate codes belonging to this domain, in numeric order
        codes = sorted(
            (c for c in output if c.startswith(letter + " ")),
            key=lambda c: [int(p) for p in c.split(" ", 1)[1].split(".")],
        )
        for code in codes:
            num = code.split(" ", 1)[1]
            levels_raw = guidance.get(code)
            text = texts.get(code)
            if not text or not levels_raw:
                skipped += 1
                continue
            sec_no = num.split(".")[0]
            sec_title = sections.get((letter, sec_no), "")
            cat_code = category_of(code)
            cat_title = texts.get(cat_code) if cat_code else None

            levels = [
                {"value": v, "label": DETAILED_LABELS.get(v, str(v)), "scenario": s}
                for v, s in sorted(levels_raw)
            ]
            rec = output[code]
            for n in rec["nist"]:
                if n in nist_desc:
                    nist_used[n] = nist_desc[n]
            q = {
                "code": code,
                "text": text,
                "category": cat_title,
                "importance": rec["importance"],
                "levels": levels,
                "nist": rec["nist"],
            }
            sec = dom_secs.setdefault(sec_no, {"no": sec_no, "title": sec_title, "questions": []})
            sec["questions"].append(q)
            kept += 1

        if dom_secs:
            ordered = [dom_secs[k] for k in sorted(dom_secs, key=lambda x: int(x))]
            domains.append({"letter": letter, "title": dom_title, "sections": ordered})

    # attach question codes onto each NIST subcategory (reverse index)
    nist_index = {}
    for code in output:
        for n in output[code]["nist"]:
            if n in nist_used:
                nist_index.setdefault(n, dict(nist_used[n], code=n, questions=[]))
                # only count questions actually kept
    # rebuild question lists from kept domains
    kept_codes = {q["code"] for d in domains for s in d["sections"] for q in s["questions"]}
    for n in nist_used:
        entry = dict(nist_used[n], code=n, questions=[])
        for code in output:
            if n in output[code]["nist"] and code in kept_codes:
                entry["questions"].append(code)
        nist_index[n] = entry

    data = {
        "meta": {
            "framework": "SOC-CMM v2.4.2 (Maturity)",
            "scale": DETAILED_LABELS,
            "domainsCount": len(domains),
            "questionsCount": kept,
        },
        "domains": domains,
        "nist": nist_index,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"kept {kept} maturity questions, skipped {skipped}")
    print(f"domains: {[ (d['title'], sum(len(s['questions']) for s in d['sections'])) for d in domains ]}")
    print(f"NIST subcategories referenced: {len(nist_index)}")
    print(f"written -> {OUT}")


if __name__ == "__main__":
    main()
